"""T-016: report runs (csv/xlsx/pdf), seeded definitions, role-scoped KPIs."""

import csv
import io

import pytest
from openpyxl import load_workbook

from apps.matching.tests.helpers import enroll_person
from apps.reports.models import ReportDefinition, ReportRun

pytestmark = pytest.mark.django_db

DEFINITIONS_URL = "/api/v1/reports/definitions/"
RUN_URL = "/api/v1/reports/run/"
RUNS_URL = "/api/v1/reports/runs/"
KPI_URL = "/api/v1/dashboard/kpis/"

SEEDED_CODES = {
    "enrollment_stats",
    "verification_outcomes",
    "case_activity",
    "duplicates",
    "clearance_issuance",
}


@pytest.fixture
def supervisor(auth_client):
    return auth_client("supervisor")


def run_report(client, code: str, format: str) -> dict:
    definition = ReportDefinition.objects.get(code=code)
    resp = client.post(
        RUN_URL,
        {"definition_id": str(definition.id), "format": format},
        format="json",
    )
    assert resp.status_code == 202, resp.json()
    return resp.json()


class TestDefinitions:
    def test_five_standard_definitions_seeded(self, supervisor):
        listed = supervisor.get(DEFINITIONS_URL).json()
        assert SEEDED_CODES <= {d["code"] for d in listed["results"]}


class TestRuns:
    def test_xlsx_opens_via_openpyxl(self, supervisor):
        enroll_person(16000)  # some data to report on
        body = run_report(supervisor, "enrollment_stats", "xlsx")
        run = ReportRun.objects.get(id=body["id"])
        assert run.status == ReportRun.Status.DONE  # eager Celery

        workbook = load_workbook(io.BytesIO(run.file.open("rb").read()))
        sheet = workbook.active
        assert sheet.title.startswith("Enrollment")
        header = [cell.value for cell in sheet[1]]
        assert header == ["Metric", "Value"]
        assert sheet.max_row > 1  # has data rows

    def test_csv_renders_with_header(self, supervisor):
        body = run_report(supervisor, "clearance_issuance", "csv")
        run = ReportRun.objects.get(id=body["id"])
        content = run.file.open("rb").read().decode("utf-8-sig")
        rows = list(csv.reader(io.StringIO(content)))
        assert rows[0] == ["Metric", "Value"]

    def test_pdf_renders(self, supervisor):
        body = run_report(supervisor, "case_activity", "pdf")
        run = ReportRun.objects.get(id=body["id"])
        assert run.file.open("rb").read().startswith(b"%PDF")

    @pytest.mark.parametrize("code", sorted(SEEDED_CODES))
    def test_every_seeded_definition_runs_clean(self, supervisor, code):
        body = run_report(supervisor, code, "csv")
        run = ReportRun.objects.get(id=body["id"])
        assert run.status == ReportRun.Status.DONE, run.error

    def test_duplicates_report_lists_dedup_hits(self, supervisor):
        from apps.enrollment.services import complete_enrollment

        enrollment_a, _ = enroll_person(16100)
        complete_enrollment(enrollment_a)
        enrollment_b, _ = enroll_person(16100)  # duplicate biometrics
        complete_enrollment(enrollment_b)

        body = run_report(supervisor, "duplicates", "csv")
        run = ReportRun.objects.get(id=body["id"])
        content = run.file.open("rb").read().decode("utf-8-sig")
        assert enrollment_a.person.person_no in content

    def test_download_done_run_is_audited(self, supervisor):
        from apps.audit.models import AuditLog

        body = run_report(supervisor, "enrollment_stats", "csv")
        resp = supervisor.get(f"{RUNS_URL}{body['id']}/download/")
        assert resp.status_code == 200
        assert AuditLog.objects.filter(
            entity="reports.ReportRun",
            entity_id=body["id"],
            action=AuditLog.Action.EXPORT,
        ).exists()

    def test_bad_format_400_unknown_definition_404(self, supervisor):
        definition = ReportDefinition.objects.get(code="enrollment_stats")
        assert (
            supervisor.post(
                RUN_URL,
                {"definition_id": str(definition.id), "format": "docx"},
                format="json",
            ).status_code
            == 400
        )
        assert (
            supervisor.post(
                RUN_URL,
                {
                    "definition_id": "00000000-0000-0000-0000-000000000000",
                    "format": "csv",
                },
                format="json",
            ).status_code
            == 404
        )

    def test_operator_cannot_run_reports(self, auth_client):
        resp = auth_client("operator").get(DEFINITIONS_URL)
        assert resp.status_code == 403

    def test_auditor_can_run_and_download(self, auth_client):
        auditor = auth_client("auditor")
        body = run_report(auditor, "verification_outcomes", "csv")
        assert auditor.get(f"{RUNS_URL}{body['id']}/download/").status_code == 200


class TestKpis:
    EXPECTED_BLOCKS = {
        "admin": {
            "enrollments",
            "applications",
            "matching",
            "certificates",
            "alerts",
            "audit",
            "verification",
        },
        "supervisor": {
            "enrollments",
            "applications",
            "matching",
            "certificates",
            "alerts",
            "audit",
            "verification",
        },
        "operator": {"enrollments", "applications", "certificates"},
        "investigator": {"enrollments", "matching", "alerts"},
        "auditor": {"audit", "verification"},
    }

    @pytest.mark.parametrize("role", sorted(EXPECTED_BLOCKS))
    def test_kpi_shape_per_role(self, auth_client, role):
        resp = auth_client(role).get(KPI_URL)
        assert resp.status_code == 200
        assert set(resp.json().keys()) == self.EXPECTED_BLOCKS[role]

    def test_kpi_values_present(self, auth_client):
        enroll_person(16200)
        body = auth_client("supervisor").get(KPI_URL).json()
        assert body["enrollments"]["today"] >= 1
        assert body["enrollments"]["week"] >= body["enrollments"]["today"]
        assert "hit_rate" in body["matching"]
        assert "open" in body["alerts"]

    def test_anonymous_401(self, api_client, db):
        assert api_client.get(KPI_URL).status_code == 401
